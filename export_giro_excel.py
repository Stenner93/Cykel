"""
Export per-stage predictions for all Giro 2026 stages to Excel.

For each stage with VeloScore data:
  - Full rider ranking by expected_pts
  - Signal breakdown (VeloScore, discipline, form)
  - Uncertainty range (optimistic / base / pessimistic)
  - Optimal MILP team highlighted
  - Actual result (where available) for comparison

Output: data/giro2026_predictions.xlsx

Usage:
    python export_giro_excel.py
"""
from __future__ import annotations
import json
import math
import sys
from pathlib import Path

ROOT = Path(__file__).parent
sys.path.insert(0, str(ROOT))

from src.predictor  import predict_all
from src.optimizer  import _solve, BUDGET
from openpyxl                  import Workbook
from openpyxl.styles           import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils            import get_column_letter
from openpyxl.formatting.rule  import ColorScaleRule, CellIsRule

DATA  = ROOT / "data"
GIRO  = DATA / "giro2026"

# ── Colours ───────────────────────────────────────────────────────────────────
C_HEADER      = "1F3864"   # dark blue header
C_HEADER_FG   = "FFFFFF"
C_CAPTAIN     = "FFD700"   # gold captain
C_TEAM_MEMBER = "D9E8FB"   # light blue — in optimal team
C_WINNER      = "C6EFCE"   # light green — actual stage winner
C_MATCH       = "E8F5E9"   # very light green — correct pick
C_MISS        = "FFEBEE"   # light red — in team but scored 0
C_TAB_EVEN    = "F5F5F5"
C_SPRINT      = "FFF9C4"
C_MOUNTAIN    = "E8F5E9"
C_TT          = "E3F2FD"
C_HILLY       = "FFF3E0"
C_COBBLED     = "F3E5F5"

STAGE_TYPE_COLOURS = {
    "sprint":   C_SPRINT,
    "mountain": C_MOUNTAIN,
    "tt":       C_TT,
    "hilly":    C_HILLY,
    "cobbled":  C_COBBLED,
}

STAGE_TYPE_LABELS = {
    "sprint":   "Sprint ⚡",
    "mountain": "Bjerg ⛰️",
    "tt":       "Enkeltstart ⏱️",
    "hilly":    "Kuperet 〰️",
    "cobbled":  "Brosten 🧱",
}

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ── Column definitions ─────────────────────────────────────────────────────────
# (key, header, width, number_format)
COLS = [
    ("rank",        "#",            5,  "0"),
    ("full_name",   "Rytter",       28, "@"),
    ("team",        "Hold",          8, "@"),
    ("price",       "Pris (M)",      9, "0.0"),
    ("expected_pts","Forv. pts",    13, "#,##0"),
    ("ci_low",      "CI lav",       13, "#,##0"),
    ("ci_high",     "CI høj",       13, "#,##0"),
    ("win_prob",    "Vind-prob",     10, "0.0%"),
    ("disc_raw",    "Disciplin",     11, "0.0"),
    ("form_score",  "Form",         9,  "0.0"),
    ("vs_signal",   "VS-signal",    11, "0.0%"),
    ("veloscore_sc","VeloScore",    11, "0.0"),
    ("reasoning",   "Begrundelse",  30, "@"),
    ("actual_pts",  "Faktiske pts", 13, "#,##0"),  # if result available
    ("in_team",     "I hold?",       8, "@"),
    ("is_captain",  "Kaptajn?",      9, "@"),
]

# ── Helpers ────────────────────────────────────────────────────────────────────

def _hdr_fill(color: str):
    return PatternFill("solid", fgColor=color)

def _cell_fill(color: str):
    return PatternFill("solid", fgColor=color)

def _norm(name: str) -> str:
    return name.strip().lower()

def _last(name: str) -> str:
    return _norm(name).split()[-1]

def _lookup_pts(full_name: str, pts_map: dict) -> int | None:
    nl = _norm(full_name)
    if nl in pts_map:
        return pts_map[nl]
    last = _last(full_name)
    for k, v in pts_map.items():
        if _last(k) == last:
            return v
    return None

def _veloscore_from_pred(pred: dict, vs_list: list[dict]) -> float | None:
    """Find VeloScore score for this rider from the raw VS predictions list."""
    name = _norm(pred.get("full_name", ""))
    last = _last(pred.get("full_name", ""))
    for v in vs_list:
        vn = _norm(v.get("rider", ""))
        if vn == name or _last(vn) == last:
            return v.get("veloscore")
    return None

def compute_ci(pred: dict) -> tuple[int, int]:
    """
    Compute a simple uncertainty range for expected_pts.

    Model: a rider's actual score is roughly Poisson-like — most likely 0
    (not in the points), occasionally high. We model uncertainty as:
      - Pessimistic (10th-pctile): 15% of expected (rider misses breakaway / crashes)
      - Optimistic  (90th-pctile): expected + sqrt(variance) * 1.5
    Not a formal statistical CI, but a directional range.
    """
    ep  = pred["expected_pts"]
    var = pred.get("variance", 0)
    ci_low  = max(0, round(ep * 0.15))
    ci_high = round(ep + math.sqrt(max(0, var)) * 1.5)
    return ci_low, ci_high


# ── Sheet builder ──────────────────────────────────────────────────────────────

def write_stage_sheet(wb: Workbook, stage_num: int, vs_entry: dict,
                      riders: list[dict], co_data: dict, pcs_form: dict,
                      weights: dict, pts_map: dict | None) -> None:
    st_type  = vs_entry["stage_type"]
    vs_preds = vs_entry["predictions"]

    sheet_name = f"S{stage_num:02d}_{st_type[:3].upper()}"
    ws = wb.create_sheet(title=sheet_name)

    # ── Tab colour ────────────────────────────────────────────────────────────
    type_colour = STAGE_TYPE_COLOURS.get(st_type, "FFFFFF")
    ws.sheet_properties.tabColor = type_colour.lstrip("#")

    # ── Run predictions ───────────────────────────────────────────────────────
    predictions = predict_all(
        riders=riders,
        stage_type=st_type,
        veloscore_data=vs_preds,
        cyclingoracle_data=co_data,
        pcs_form_data=pcs_form,
        weights=weights,
    )

    # ── Run MILP for optimal team ─────────────────────────────────────────────
    team_result = _solve(predictions, budget=BUDGET, label="export")
    team_ids    = set()
    captain_id  = None
    if team_result:
        team_ids   = {r["rider_id"] for r in team_result["team"]}
        captain_id = next(
            (r["rider_id"] for r in team_result["team"] if r.get("is_captain")),
            None
        )

    # ── Stage header ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:P1")
    hdr_cell = ws["A1"]
    hdr_cell.value = (
        f"Etape {stage_num} — {STAGE_TYPE_LABELS.get(st_type, st_type.upper())}  "
        f"| VeloScore: {len(vs_preds)} ryttere  "
        + (f"| Faktiske resultater tilgængelige" if pts_map else "| Ingen faktiske resultater")
    )
    hdr_cell.font      = Font(bold=True, size=13, color=C_HEADER_FG)
    hdr_cell.fill      = _hdr_fill(C_HEADER)
    hdr_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # ── Column headers ────────────────────────────────────────────────────────
    for col_idx, (_, header, width, _fmt) in enumerate(COLS, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font      = Font(bold=True, color=C_HEADER_FG, size=10)
        cell.fill      = _hdr_fill(C_HEADER)
        cell.alignment = Alignment(horizontal="center")
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 16

    # ── Data rows ─────────────────────────────────────────────────────────────
    pred_by_id = {p["rider_id"]: p for p in predictions}

    for row_idx, pred in enumerate(predictions[:195], 3):
        ci_low, ci_high = compute_ci(pred)
        vs_sc           = _veloscore_from_pred(pred, vs_preds)
        actual_pts      = _lookup_pts(pred["full_name"], pts_map) if pts_map else None
        in_team         = pred["rider_id"] in team_ids
        is_cap          = pred["rider_id"] == captain_id

        row_data = {
            "rank":        row_idx - 2,
            "full_name":   pred["full_name"],
            "team":        pred["team"],
            "price":       pred["price"],
            "expected_pts": pred["expected_pts"],
            "ci_low":      ci_low,
            "ci_high":     ci_high,
            "win_prob":    pred["win_prob"],
            "disc_raw":    pred.get("disc_raw", 0),
            "form_score":  pred.get("form_score", 0),
            "vs_signal":   pred["signal_scores"].get("veloscore", 0),
            "veloscore_sc": vs_sc if vs_sc is not None else "",
            "reasoning":   pred.get("reasoning", ""),
            "actual_pts":  actual_pts if actual_pts is not None else "",
            "in_team":     "✓" if in_team else "",
            "is_captain":  "★" if is_cap else "",
        }

        for col_idx, (key, _, _, fmt) in enumerate(COLS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data[key])
            cell.border    = BORDER
            cell.alignment = Alignment(vertical="center")
            if fmt not in ("@",):
                cell.number_format = fmt

        # ── Row colouring ─────────────────────────────────────────────────────
        if is_cap:
            fill = _cell_fill(C_CAPTAIN)
        elif in_team:
            fill = _cell_fill(C_TEAM_MEMBER)
        elif actual_pts and actual_pts > 0 and row_idx - 2 <= 5:
            fill = _cell_fill(C_WINNER)
        elif row_idx % 2 == 0:
            fill = _cell_fill(C_TAB_EVEN)
        else:
            fill = None

        if fill:
            for col_idx in range(1, len(COLS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    # ── Freeze panes ─────────────────────────────────────────────────────────
    ws.freeze_panes = "A3"

    # ── Conditional formatting: expected_pts gradient ────────────────────────
    pts_col = get_column_letter(5)  # expected_pts is col 5
    ws.conditional_formatting.add(
        f"{pts_col}3:{pts_col}{len(predictions)+2}",
        ColorScaleRule(
            start_type="min",  start_color="FFFFFF",
            mid_type="percentile", mid_value=75, mid_color="FFEB84",
            end_type="max",    end_color="F8696B",
        ),
    )


# ── Summary sheet ──────────────────────────────────────────────────────────────

def write_summary_sheet(wb: Workbook, stage_summaries: list[dict]) -> None:
    ws = wb.create_sheet(title="Oversigt", index=0)

    # Header
    ws.merge_cells("A1:L1")
    c = ws["A1"]
    c.value = "Giro 2026 — Modeloversigt (alle etaper med VeloScore)"
    c.font      = Font(bold=True, size=14, color=C_HEADER_FG)
    c.fill      = _hdr_fill(C_HEADER)
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 24

    sum_headers = [
        ("#",          4),
        ("Etape",      7),
        ("VS-type",   10),
        ("Kaptajn",   25),
        ("Top-3 picks",35),
        ("Forv. pts", 13),
        ("Faktiske pts (model)", 22),
        ("Oracle pts",13),
        ("Efficiency", 11),
        ("Vinder?",   25),
        ("Vinder i hold?", 15),
        ("Type match?", 13),
    ]
    for col_idx, (hdr, width) in enumerate(sum_headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=hdr)
        cell.font      = Font(bold=True, color=C_HEADER_FG, size=10)
        cell.fill      = _hdr_fill(C_HEADER)
        cell.alignment = Alignment(horizontal="center")
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, s in enumerate(stage_summaries, 3):
        vals = [
            row_idx - 2,
            s["stage"],
            s["vs_type"],
            s.get("captain_name", ""),
            s.get("top3", ""),
            s.get("model_pred_pts", ""),
            s.get("model_actual_pts", ""),
            s.get("oracle_pts", ""),
            s.get("efficiency"),
            s.get("actual_winner", ""),
            "Ja ★" if s.get("winner_in_model") else "Nej",
            "✓" if s.get("type_match") else "✗",
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border    = BORDER
            cell.alignment = Alignment(vertical="center")
            if col_idx == 9 and val is not None:   # efficiency
                cell.number_format = "0.0%"
            if col_idx in (6, 7, 8):
                cell.number_format = "#,##0"

        # Colour by type match
        type_match = s.get("type_match", False)
        row_fill = _cell_fill("E8F5E9") if type_match else _cell_fill("FFEBEE")
        for col_idx in range(1, len(sum_headers) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = row_fill

    ws.freeze_panes = "A3"


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    print("Indlæser data…")
    riders  = json.loads((DATA / "riders.json").read_text(encoding="utf-8"))
    weights = json.loads((DATA / "calibrated_weights.json").read_text(encoding="utf-8"))

    vs_raw  = json.loads((GIRO / "veloscore.json").read_text(encoding="utf-8"))
    vs_data = {s["stage"]: s for s in vs_raw}

    res_raw = json.loads((GIRO / "results.json").read_text(encoding="utf-8"))
    res_data = {
        s["stage"]: {
            "pts_map": {
                r["rider"].strip().lower(): int(r["points"])
                for r in s.get("top_results", [])
                if r.get("points") is not None
            },
            "winner": next(
                (r["rider"] for r in s.get("top_results", []) if r.get("rank") == 1),
                ""
            ),
            "stage_type": s.get("stage_type", ""),
        }
        for s in res_raw
        if s.get("top_results")
    }

    # Load cache data
    co_path = DATA / "cache" / "cyclingoracle.json"
    if co_path.exists():
        co_raw  = json.loads(co_path.read_text(encoding="utf-8"))
        co_data = {k: v.get("ratings", {}) for k, v in co_raw.items()}
    else:
        co_data = {}

    pcs_path = DATA / "cache" / "pcs_form.json"
    if pcs_path.exists():
        pcs_raw = json.loads(pcs_path.read_text(encoding="utf-8"))
        pcs_form = {}
        for rid, v in pcs_raw.items():
            if v.get("not_found"):
                continue
            if "form_by_type" in v:
                pcs_form[rid] = v["form_by_type"]
            else:
                pcs_form[rid] = {"overall": v.get("form_score", 50.0)}
    else:
        pcs_form = {}

    print(f"  Ryttere: {len(riders)}  |  VeloScore stages: {sorted(vs_data)}  |  CO: {len(co_data)}  |  PCS form: {len(pcs_form)}")

    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    stage_summaries = []

    # Also import retro_optimizer helpers for oracle scoring
    import sys as _sys
    from retro_optimizer import (
        build_oracle_preds, score_team, team_rows, _lookup_pts as _retro_lookup
    )

    for stage_num in sorted(vs_data.keys()):
        vs_entry = vs_data[stage_num]
        pts_map  = res_data.get(stage_num, {}).get("pts_map")
        print(f"  Etape {stage_num:2d} ({vs_entry['stage_type']:9s})…", end=" ", flush=True)

        predictions = predict_all(
            riders=riders,
            stage_type=vs_entry["stage_type"],
            veloscore_data=vs_entry["predictions"],
            cyclingoracle_data=co_data,
            pcs_form_data=pcs_form,
            weights=weights,
        )

        # Optimal team
        team_result = _solve(predictions, budget=BUDGET, label="export")
        team_ids    = set()
        captain_id  = None
        captain_name = ""
        if team_result:
            team_ids     = {r["rider_id"] for r in team_result["team"]}
            captain      = next((r for r in team_result["team"] if r.get("is_captain")), None)
            captain_id   = captain["rider_id"] if captain else None
            captain_name = captain["full_name"] if captain else ""

        top3 = ", ".join(p["full_name"] for p in predictions[:3])

        # Oracle scoring (if results available)
        oracle_pts_val   = None
        model_actual_val = None
        efficiency       = None
        winner_in_model  = False
        actual_winner    = res_data.get(stage_num, {}).get("winner", "")
        type_match       = (
            vs_entry["stage_type"] == res_data.get(stage_num, {}).get("stage_type", "")
            if stage_num in res_data else None
        )

        if pts_map:
            oracle_preds  = build_oracle_preds(riders, pts_map)
            oracle_result = _solve(oracle_preds, budget=BUDGET, label="oracle")
            if oracle_result:
                oracle_sc       = score_team(oracle_result, pts_map)
                oracle_pts_val  = oracle_sc["total_pts"]
            if team_result:
                model_sc        = score_team(team_result, pts_map)
                model_actual_val = model_sc["total_pts"]
            if oracle_pts_val and oracle_pts_val > 0 and model_actual_val is not None:
                efficiency = model_actual_val / oracle_pts_val
            if team_result and actual_winner:
                winner_in_model = any(
                    _last(actual_winner) in _norm(r["full_name"])
                    for r in team_result["team"]
                )

        eff_str = f"{efficiency*100:.1f}%" if efficiency is not None else "N/A"
        print(f"oracle={oracle_pts_val or 0:>10,.0f}  model={model_actual_val or 0:>10,.0f}  eff={eff_str}")

        stage_summaries.append({
            "stage":           stage_num,
            "vs_type":         vs_entry["stage_type"],
            "captain_name":    captain_name,
            "top3":            top3,
            "model_pred_pts":  team_result["expected_pts"] if team_result else 0,
            "model_actual_pts": model_actual_val,
            "oracle_pts":       oracle_pts_val,
            "efficiency":       efficiency,
            "actual_winner":    actual_winner,
            "winner_in_model":  winner_in_model,
            "type_match":       type_match,
        })

        # Write per-stage sheet
        write_stage_sheet(
            wb, stage_num, vs_entry, riders, co_data, pcs_form, weights, pts_map
        )

    write_summary_sheet(wb, stage_summaries)

    out_path = DATA / "giro2026_predictions.xlsx"
    wb.save(str(out_path))
    print(f"\n  ✓ Excel gemt: {out_path}")

    # ── Summary print ──────────────────────────────────────────────────────────
    valid  = [s for s in stage_summaries if s["efficiency"] is not None]
    typed  = [s for s in valid if s["type_match"]]
    wim    = sum(1 for s in valid if s["winner_in_model"])
    avg_e  = sum(s["efficiency"] for s in valid) / len(valid) if valid else 0
    avg_em = sum(s["efficiency"] for s in typed) / len(typed) if typed else 0

    print(f"\n  === Sammenfatning ===")
    print(f"  Etaper analyseret: {len(stage_summaries)}  ({len(valid)} med faktiske resultater)")
    print(f"  Gns. efficiency (alle):       {avg_e*100:.1f}%")
    print(f"  Gns. efficiency (type-match): {avg_em*100:.1f}%")
    print(f"  Vinder i modelholdets hold:   {wim}/{len(valid)} ({wim/len(valid)*100:.0f}%)" if valid else "")


if __name__ == "__main__":
    main()
